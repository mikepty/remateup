"""
Endpoint de exportación a Excel.
Genera archivos .xlsx con el mismo formato que usa el cliente
(FEED MAIN PANAMA.xlsx / REMATEHOY FEED MAIN COLOMBIAv.xlsx).
"""
import io
from datetime import datetime
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ..database import get_db
from ..models import Aviso

router = APIRouter(prefix="/exportar", tags=["exportar"])

# Estilos
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="2D2B55", end_color="2D2B55", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def _estilo_header(ws, columns):
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _estilo_celda(cell):
    cell.alignment = CELL_ALIGN
    cell.border = THIN_BORDER


def _formatear_hora(hora):
    """Convierte hora a formato HH:MM para el Excel."""
    if not hora:
        return None
    if isinstance(hora, str):
        return hora
    return hora.strftime("%H:%M")


def _generar_excel_panama(avisos):
    wb = Workbook()
    ws = wb.active
    ws.title = "HOJA PA"

    columnas = [
        None, "pais", "codigo", "expediente", "juzg lugar", "observaciones",
        "proceso", "demandante", "demandado", "fecha", "hora",
        "codigo ubicacion", "finca/matr", "plano", "lote/casa", "superficie",
        "categoria", "descripcion", "provincia", "base", "fianza", "minimo",
        "cod prensa"
    ]

    _estilo_header(ws, columnas)

    for row_idx, a in enumerate(avisos, 2):
        data = [
            None,  # columna vacía inicial
            a.pais,
            a.codigo or f"PA{a.id:08d}",
            a.expediente,
            a.lugar,
            None,  # observaciones
            a.proceso,
            a.demandante,
            a.demandado,
            a.fecha,
            _formatear_hora(a.hora),
            a.codigo_ubicacion,
            a.finca_matr,
            a.plano,
            a.lote_casa,
            a.superficie,
            a.categoria_codigo,
            a.descripcion_completa or a.descripcion,
            a.provincia,
            float(a.base) if a.base else None,
            float(a.fianza) if a.fianza else None,
            float(a.minimo) if a.minimo else None,
            a.codigo_fuente,
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _estilo_celda(cell)

    # Ajustar anchos de columna
    anchos = {2: 8, 3: 16, 4: 16, 5: 30, 7: 15, 8: 30, 9: 30, 10: 12, 11: 8,
              12: 12, 13: 12, 14: 10, 15: 15, 16: 12, 17: 10, 18: 50, 19: 15,
              20: 15, 21: 12, 22: 12, 23: 15}
    for col, width in anchos.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.freeze_panes = "A2"
    return wb


def _generar_excel_colombia(avisos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    columnas = [
        None, "pais", "codigo", "fecha", "hora", "proceso", "expediente",
        "juzg lugar", "observaciones", "categoria", "demandante", "demandado",
        "lote/casa", "descripcion", "superficie", "finca/matr",
        "codigo ubicacion", "provincia", "plano", "base", "fianza", "minimo",
        "correo", "cod prensa"
    ]

    _estilo_header(ws, columnas)

    for row_idx, a in enumerate(avisos, 2):
        data = [
            None,
            a.pais,
            a.codigo or f"CO{a.id:08d}",
            a.fecha,
            _formatear_hora(a.hora),
            a.proceso,
            a.expediente,
            a.lugar,
            None,
            a.categoria_codigo,
            a.demandante,
            a.demandado,
            a.lote_casa,
            a.descripcion_completa or a.descripcion,
            a.superficie,
            a.finca_matr,
            a.codigo_ubicacion,
            a.provincia,
            a.plano,
            float(a.base) if a.base else None,
            40,  # fianza siempre 40% en Colombia
            100,  # minimo porcentaje
            None,  # correo
            a.codigo_fuente,
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _estilo_celda(cell)

    anchos = {2: 8, 3: 16, 4: 12, 5: 8, 6: 15, 7: 20, 8: 25, 10: 10,
              11: 30, 12: 30, 13: 15, 14: 50, 15: 12, 16: 15, 17: 12,
              18: 15, 20: 15, 21: 10, 22: 10, 23: 40, 24: 15}
    for col, width in anchos.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.freeze_panes = "A2"
    return wb


def _workbook_a_stream(wb):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/excel")
def exportar_excel(
    pais: str = Query(None, description="PA, CO, o None para todos"),
    estado: str = Query(None, description="subido, auto_aprobado, etc."),
    ids: str = Query(None, description="IDs separados por coma, ej: 1,2,3"),
    db: Session = Depends(get_db),
):
    """Exporta avisos a Excel. Soporta filtros: pais, estado, o IDs especificos."""
    query = db.query(Aviso).filter(Aviso.estado != "eliminado")

    if pais:
        pais_code = 1 if pais.upper() == "PA" else 2
        query = query.filter(Aviso.pais == pais_code)

    if estado:
        query = query.filter(Aviso.estado == estado)

    if ids:
        id_list = [int(i.strip()) for i in ids.split(",") if i.strip().isdigit()]
        if id_list:
            query = query.filter(Aviso.id.in_(id_list))

    avisos = query.order_by(Aviso.id).all()

    if not avisos:
        return {"message": "No hay avisos para exportar"}

    # Separar por país si hay mixto
    pa = [a for a in avisos if a.pais == 1]
    co = [a for a in avisos if a.pais == 2]

    if pais and pais.upper() == "PA" and pa:
        wb = _generar_excel_panama(pa)
        filename = f"FEED_MAIN_PANAMA_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    elif pais and pais.upper() == "CO" and co:
        wb = _generar_excel_colombia(co)
        filename = f"FEED_MAIN_COLOMBIA_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    else:
        # Ambos países en un solo archivo con dos hojas
        wb = Workbook()
        wb.remove(wb.active)

        if pa:
            ws_pa = wb.create_sheet("HOJA PA")
            # Reusar generador
            wb_pa = _generar_excel_panama(pa)
            ws_pa_data = wb_pa.active
            for row in ws_pa_data.iter_rows():
                for cell in row:
                    new_cell = ws_pa.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.alignment = cell.alignment.copy()
                        new_cell.border = cell.border.copy()

        if co:
            ws_co = wb.create_sheet("HOJA CO")
            wb_co = _generar_excel_colombia(co)
            ws_co_data = wb_co.active
            for row in ws_co_data.iter_rows():
                for cell in row:
                    new_cell = ws_co.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.alignment = cell.alignment.copy()
                        new_cell.border = cell.border.copy()

        filename = f"FEED_MAIN_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    buffer = _workbook_a_stream(wb)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/resumen")
def resumen_export(
    pais: str = Query(None),
    db: Session = Depends(get_db),
):
    """Devuelve un resumen de lo que se exportaría."""
    query = db.query(Aviso).filter(Aviso.estado != "eliminado")
    if pais:
        pais_code = 1 if pais.upper() == "PA" else 2
        query = query.filter(Aviso.pais == pais_code)

    total = query.count()
    subidos = query.filter(Aviso.estado == "subido").count()
    auto = query.filter(Aviso.estado == "auto_aprobado").count()
    pendientes = query.filter(Aviso.estado == "esperando_aprobacion").count()

    return {
        "total": total,
        "subidos": subidos,
        "auto_aprobados": auto,
        "pendientes": pendientes,
        "listos_para_exportar": subidos + auto,
    }
